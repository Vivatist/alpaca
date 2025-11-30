#!/usr/bin/env python3
"""Excel parser capable of handling legacy XLS and large multi-sheet workbooks."""

from __future__ import annotations

import os
import re
import shutil
from datetime import date, datetime, time
from pathlib import Path
from typing import TYPE_CHECKING, Iterable, List, Optional, Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

try:  # xlrd нужен только для наследия .xls, поэтому импортируем лениво
    import xlrd  # type: ignore
except Exception:  # pragma: no cover
    xlrd = None  # type: ignore

from ..base_parser import BaseParser
from ..document_converter import convert_xls_to_xlsx

if TYPE_CHECKING:  # pragma: no cover
    from alpaca.domain.files.models import FileSnapshot


class ExcelParser(BaseParser):
    """Парсер Excel документов (.xlsx/.xls) с поддержкой сложных таблиц."""

    def __init__(self, max_rows_per_table: int = 200):
        super().__init__("excel-parser")
        self.max_rows_per_table = max_rows_per_table

    def _parse(self, file: "FileSnapshot") -> str:
        file_path = file.full_path
        suffix = Path(file_path).suffix.lower()
        working_path = file_path
        cleanup_dir: Optional[Path] = None

        try:
            if not os.path.exists(file_path):
                self.logger.error(f"File not found | file={file.path}")
                raise FileNotFoundError(f"File not found | file={file.path}")

            if suffix == ".xls":
                converted = convert_xls_to_xlsx(file_path)
                if not converted:
                    self.logger.warning(
                        "LibreOffice conversion failed, falling back to xlrd for .xls parsing"
                    )
                    xlrd_text = self._parse_xls_with_xlrd(file_path)
                    return xlrd_text
                working_path = converted
                cleanup_dir = Path(converted).parent
                self.logger.info("Converted legacy .xls to .xlsx for parsing")
            elif suffix != ".xlsx":
                self.logger.error(f"Unsupported Excel extension | ext={suffix}")
                raise ValueError(f"Unsupported Excel extension | ext={suffix}")

            try:
                workbook = load_workbook(working_path, data_only=True, read_only=True)
            except Exception as exc:
                self.logger.error(f"Failed to load workbook | file={file.path} error={exc}")
                raise RuntimeError(
                    f"Failed to load Excel workbook | file={file.path}"
                ) from exc

            sheet_blocks: List[str] = []
            for sheet in workbook.worksheets:
                sheet_md = self._sheet_to_markdown(sheet)
                if sheet_md:
                    sheet_blocks.append(f"## Лист: {sheet.title or 'Sheet'}\n\n{sheet_md}")

            result = "\n\n---\n\n".join(sheet_blocks).strip()
            self.logger.info(
                f"Excel parsing complete | sheets={len(sheet_blocks)} length={len(result)}"
            )
            return result

        finally:
            if cleanup_dir and cleanup_dir.exists():
                shutil.rmtree(cleanup_dir, ignore_errors=True)

    def _sheet_to_markdown(self, sheet: Worksheet) -> str:
        rows = [
            [self._format_cell(value) for value in row]
            for row in sheet.iter_rows(values_only=True)
        ]
        return self._rows_to_markdown(rows)

    def _rows_to_markdown(self, rows: List[List[str]]) -> str:
        rows = [row for row in rows if any(cell for cell in row)]
        if not rows:
            return ""

        header_idx = self._detect_header_row(rows)
        preface_rows = rows[:header_idx]
        table_rows = rows[header_idx:]
        if not table_rows:
            table_rows = rows

        header = table_rows[0]
        data_rows = table_rows[1:]

        parts: List[str] = []
        preface_text = self._rows_to_paragraphs(preface_rows)
        if preface_text:
            parts.append(preface_text)

        table_text = self._render_table(header, data_rows)
        if table_text:
            parts.append(table_text)

        return "\n\n".join(parts).strip()

    def _render_table(self, header: List[str], rows: List[List[str]]) -> str:
        max_cols = max(len(header), *(len(row) for row in rows)) if rows else len(header)
        max_cols = max(max_cols, 1)
        header = self._pad_row(header, max_cols)
        row_chunks = list(self._chunk_rows(rows, self.max_rows_per_table)) or [[]]

        tables: List[str] = []
        for idx, chunk in enumerate(row_chunks, start=1):
            padded_chunk = [self._pad_row(row, max_cols) for row in chunk]
            lines = ["| " + " | ".join(header) + " |"]
            lines.append("|" + "|".join(["---"] * max_cols) + "|")
            for row in padded_chunk:
                lines.append("| " + " | ".join(row) + " |")
            table_block = "\n".join(lines)
            if len(row_chunks) > 1:
                tables.append(f"Табличный блок {idx}/{len(row_chunks)}\n\n{table_block}")
            else:
                tables.append(table_block)
        return "\n\n".join(tables)

    def _pad_row(self, row: List[str], max_cols: int) -> List[str]:
        return row + ["" for _ in range(max_cols - len(row))]

    def _rows_to_paragraphs(self, rows: List[List[str]]) -> str:
        sentences = []
        for row in rows:
            sentence = " ".join(cell for cell in row if cell).strip()
            if sentence:
                sentences.append(sentence)
        return "\n".join(sentences)

    def _detect_header_row(self, rows: List[List[str]]) -> int:
        best_idx = 0
        best_score = -1
        for idx, row in enumerate(rows):
            filled = sum(1 for cell in row if cell)
            if filled == 0:
                continue
            score = filled * 10 - idx  # предпочитаем более ранние строки
            if score > best_score:
                best_idx = idx
                best_score = score
            if filled >= 3:
                break
        return best_idx

    def _chunk_rows(
        self, rows: List[List[str]], chunk_size: int
    ) -> Iterable[List[List[str]]]:
        if chunk_size <= 0:
            yield rows
            return
        for start in range(0, len(rows), chunk_size):
            yield rows[start : start + chunk_size]

    def _format_cell(self, value: object) -> str:
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d %H:%M:%S").rstrip(" 00:00:00")
        if isinstance(value, date):
            return value.strftime("%Y-%m-%d")
        if isinstance(value, time):
            return value.strftime("%H:%M:%S")
        if isinstance(value, float):
            return ("%.6f" % value).rstrip("0").rstrip(".")
        return self._normalize_cell_text(str(value))

    def _normalize_cell_text(self, text: str) -> str:
        if not text:
            return ""
        collapsed = re.sub(r"[\r\n]+", " ", text)
        collapsed = re.sub(r"\s+", " ", collapsed)
        return collapsed.strip()

    def _parse_xls_with_xlrd(self, file_path: str) -> str:
        if xlrd is None:  # pragma: no cover
            self.logger.error("xlrd is not installed, cannot parse legacy .xls file")
            return ""

        try:
            workbook = xlrd.open_workbook(file_path, formatting_info=False)
        except Exception as exc:
            self.logger.error(f"Failed to open .xls with xlrd | file={file_path} error={exc}")
            return ""

        try:
            sheet_blocks: List[str] = []
            for sheet in workbook.sheets():
                rows: List[List[str]] = []
                for row_idx in range(sheet.nrows):
                    row_values: List[str] = []
                    for col_idx in range(sheet.ncols):
                        cell_value = sheet.cell_value(row_idx, col_idx)
                        cell_type = sheet.cell_type(row_idx, col_idx)
                        row_values.append(
                            self._format_xlrd_cell(cell_value, cell_type, workbook)
                        )
                    rows.append(row_values)

                sheet_md = self._rows_to_markdown(rows)
                if sheet_md:
                    sheet_blocks.append(f"## Лист: {sheet.name or 'Sheet'}\n\n{sheet_md}")

            result = "\n\n---\n\n".join(sheet_blocks).strip()
            self.logger.info(
                f"Excel parsing via xlrd complete | sheets={len(sheet_blocks)} length={len(result)}"
            )
            return result
        finally:
            try:
                workbook.release_resources()
            except Exception:  # pragma: no cover
                pass

    def _format_xlrd_cell(self, value: object, cell_type: int, workbook: Any) -> str:
        if xlrd is None:
            return self._format_cell(value)

        if cell_type == xlrd.XL_CELL_EMPTY:
            return ""
        if cell_type == xlrd.XL_CELL_BOOLEAN:
            return "TRUE" if bool(value) else "FALSE"
        if cell_type == xlrd.XL_CELL_ERROR:
            return "#ERROR"
        if cell_type == xlrd.XL_CELL_DATE:
            try:
                y, m, d, hh, mm, ss = xlrd.xldate_as_tuple(value, workbook.datemode)
                if (y, m, d) == (0, 0, 0):
                    return self._format_cell(time(hh, mm, ss))
                if (hh, mm, ss) == (0, 0, 0):
                    return self._format_cell(date(y, m, d))
                return self._format_cell(datetime(y, m, d, hh, mm, ss))
            except Exception:
                return self._format_cell(value)
        return self._format_cell(value)
